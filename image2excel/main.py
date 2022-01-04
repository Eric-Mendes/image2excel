import os
import urllib.request
from configparser import ConfigParser

import validators
import numpy as np
import pandas as pd
from openpyxl import load_workbook, styles, utils
from PIL import Image

if __name__ == "__main__":
    config_parser = ConfigParser()
    config_parser.read("image2excel/config.ini")
    default_configs = config_parser["DEFAULT"]

    # Loading the image as a PIL Image
    image_name = default_configs["image"]
    if validators.url(image_name):
        urllib.request.urlretrieve(
            image_name, 
            f"images/{os.path.split(image_name)[1]}"
        )
        image_name = os.path.split(image_name)[1]

    img = Image.open(f"images/{image_name}").convert("RGB")

    # Resizing it and getting it as a 2d list with the RGB colors of each pixel
    FACTOR = int(default_configs["factor"])
    img = img.resize((img.size[0] // FACTOR, img.size[1] // FACTOR))
    arr = np.array(img).tolist()
    # OpenPyxl colors work in a weird way
    arr = [["%02x%02x%02x" % tuple(item) for item in row] for row in arr]

    df = pd.DataFrame(arr)
    # the same that was defined above but without extension
    image_name = os.path.splitext(image_name)[0]

    # Saving a DataFrame where each cell has a text corresponding to the RGB color its background should be
    df.to_excel(f"spreadsheets/{image_name}.xlsx", index=False, header=False)

    # Loading the excel file, painting each cell with its color and saving the updates
    wb = load_workbook(f"spreadsheets/{image_name}.xlsx")

    ws = wb.active
    ws.title = image_name

    for row in range(1, df.shape[0] + 1):
        for col in range(1, df.shape[1] + 1):
            cell = ws.cell(row=row, column=col)
            # Attempt to make cell a square
            # You can tweak these numbers to make the image better for you
            ws.row_dimensions[row].height = float(default_configs["row_height"])
            ws.column_dimensions[utils.get_column_letter(col)].width = float(
                default_configs["column_width"]
            )

            # Painting the cell
            cell.fill = styles.PatternFill(
                start_color=cell.value,
                end_color=cell.value,
                fill_type="solid"
            )
            if not default_configs.getboolean("cell_value"):
                cell.value = None  # Deletes the text from the cell

    # Saves spreadsheet already zoomed out
    ws.sheet_view.zoomScale = int(default_configs["zoom_scale"])
    wb.save(f"spreadsheets/{image_name}.xlsx")
