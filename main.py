from openpyxl import (Workbook, load_workbook, styles)
from openpyxl.styles.colors import Color
import pandas as pd
from PIL import Image
from PIL import ImageOps
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors

img = Image.open(r'./image.png').convert("RGB")
img = img.resize((img.size[0]//6, img.size[1]//2))
arr = np.array(img).tolist()
arr = [['%02x%02x%02x' % tuple(item) for item in row] for row in arr]

df = pd.DataFrame(arr)
df.to_excel("test.xlsx", index=False, header=False)

wb = load_workbook('test.xlsx')

ws = wb.active
ws.title = "test"

for row in range(1, df.shape[0]+1):
    for col in range(1, df.shape[1]+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = styles.PatternFill(start_color=cell.value, end_color=cell.value, fill_type="solid")

wb.save('test2.xlsx')
